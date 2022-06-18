from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import datas
zones_test = ('defence', 'middle', 'atack', 'shoots', 'other')
first_half = {'defence': {'Сопр-назад': 1, 'Сопр-назад❌': 1, 'Сопр-поперек': 0, 'Сопр-поперек❌': 0, 'Сопр-вперед': 0,
                          'Сопр-вперед❌': 0, 'Назад': 0, 'Назад❌': 0, 'Поперек': 0, 'Поперек❌': 0, 'Вперед': 0,
                          'Вперед❌': 0, 'Обводка✅': 0, 'Обводка❌': 7, 'Отбор✅': 0, 'Отбор❌': 0, 'Перехват✅': 555,
                          'Перехват❌': 0},
              'middle': {'Сопр-назад': 0, 'Сопр-назад❌': 0, 'Сопр-поперек': 0, 'Сопр-поперек❌': 0, 'Сопр-вперед': 0,
                         'Сопр-вперед❌': 0, 'Назад': 0, 'Назад❌': 0, 'Поперек': 0, 'Поперек❌': 0, 'Вперед': 0,
                         'Вперед❌': 0, 'Обводка✅': 0, 'Обводка❌': 0, 'Отбор✅': 0, 'Отбор❌': 0, 'Перехват✅': 0,
                         'Перехват❌': 0},
              'atack': {'Сопр-назад': 0, 'Сопр-назад❌': 0, 'Сопр-поперек': 0, 'Сопр-поперек❌': 0, 'Сопр-вперед': 0,
                        'Сопр-вперед❌': 0, 'Назад': 0, 'Назад❌': 0, 'Поперек': 0, 'Поперек❌': 0, 'Вперед': 0,
                        'Вперед❌': 0, 'Обводка✅': 0, 'Обводка❌': 0, 'Отбор✅': 0, 'Отбор❌': 0, 'Перехват✅': 0,
                        'Перехват❌': 10},
              'shoots': {'из-за штрафной': 0, 'из-за штрафной❌': 0, 'ИЗ штрафной': 0, 'ИЗ штрафной❌': 0},
              'other': {'Потери мяча': 0, 'Подбор мяча': 0, 'Угловые': 0, 'Штрафные': 0, 'Прием мяча': 0,
                        'Прием мяча❌': 0}}
second_half = {'defence': {'Сопр-назад': 0, 'Сопр-назад❌': 0, 'Сопр-поперек': 0, 'Сопр-поперек❌': 0, 'Сопр-вперед': 0,
                           'Сопр-вперед❌': 0, 'Назад': 0, 'Назад❌': 0, 'Поперек': 0, 'Поперек❌': 0, 'Вперед': 0,
                           'Вперед❌': 0, 'Обводка✅': 0, 'Обводка❌': 0, 'Отбор✅': 0, 'Отбор❌': 0, 'Перехват✅': 0,
                           'Перехват❌': 0},
               'middle': {'Сопр-назад': 0, 'Сопр-назад❌': 0, 'Сопр-поперек': 0, 'Сопр-поперек❌': 0, 'Сопр-вперед': 1,
                          'Сопр-вперед❌': 1, 'Назад': 0, 'Назад❌': 0, 'Поперек': 0, 'Поперек❌': 0, 'Вперед': 0,
                          'Вперед❌': 0, 'Обводка✅': 0, 'Обводка❌': 0, 'Отбор✅': 0, 'Отбор❌': 0, 'Перехват✅': 0,
                          'Перехват❌': 0},
               'atack': {'Сопр-назад': 0, 'Сопр-назад❌': 0, 'Сопр-поперек': 1, 'Сопр-поперек❌': 1, 'Сопр-вперед': 0,
                         'Сопр-вперед❌': 0, 'Назад': 0, 'Назад❌': 0, 'Поперек': 0, 'Поперек❌': 0, 'Вперед': 0,
                         'Вперед❌': 0, 'Обводка✅': 0, 'Обводка❌': 0, 'Отбор✅': 0, 'Отбор❌': 0, 'Перехват✅': 0,
                         'Перехват❌': 0},
               'shoots': {'из-за штрафной': 0, 'из-за штрафной❌': 0, 'ИЗ штрафной': 0, 'ИЗ штрафной❌': 0},
               'other': {'Потери мяча': 0, 'Подбор мяча': 15, 'Угловые': 11, 'Штрафные': 10, 'Прием мяча': 7,
                         'Прием мяча❌': 4}}

ACTIONS_FOR_PERSONAL = ('Сопр-назад',
                        'Сопр-назад❌',
                        'Сопр-поперек',
                        'Сопр-поперек❌',
                        'Сопр-вперед',
                        'Сопр-вперед❌',
                        'Назад',
                        'Назад❌',
                        'Поперек',
                        'Поперек❌',
                        'Вперед',
                        'Вперед❌',
                        'Обводка✅',
                        'Обводка❌',
                        'Отбор✅',
                        'Отбор❌',
                        'Перехват✅',
                        'Перехват❌')


def _get_result_activities_with_zones(data: dict, check_list, ws: Workbook.active):
    col_title = 4
    row_title = 11
    for zones in check_list:
        n = 0
        for activity in ACTIONS_FOR_PERSONAL[::2]:
            ws.cell(row=row_title, column=col_title, value=data[zones][activity])
            n += 1
            col_title += 1
        col_title = col_title - n
        for activity in ACTIONS_FOR_PERSONAL[1::2]:
            ws.cell(row=row_title + 1, column=col_title, value=data[zones][activity])
            col_title += 1


def _get_shoots_results(data:dict, ws: Workbook.active):
    col_title = 31
    row_title = 11
    for activity in datas.SHOOTS[::2]:
        ws.cell(row=row_title, column=col_title, value=data[activity])
        col_title += 1
    col_title = 31
    for activity in datas.SHOOTS[1::2]:
        ws.cell(row=row_title + 1, column=col_title, value=data[activity])
        col_title +=1

def _get_other_results(data: dict, ws: Workbook.active):
    row_title =  11
    col_title = 33
    for activity in datas.OTHER[:-2]:
        ws.cell(row=row_title, column=col_title, value=data[activity])
        col_title += 1
    for activity in datas.OTHER[-2:]:
        ws.cell(row=row_title, column=col_title, value=data[activity])
        row_title += 1



def _prepare_half(dict_data: dict, ws: Workbook.active):
    _get_result_activities_with_zones(dict_data,zones_test[:3], ws)
    _get_shoots_results(dict_data['shoots'], ws)
    _get_other_results(dict_data['other'], ws)

def prepare_workbook():
    wb = load_workbook("Персональная статистика 2.0 - 4.xlsx")
    sh = wb.sheetnames[:2]
    first_half_sheet = wb[sh[0]]
    second_half_sheet = wb[sh[1]]
    _prepare_half(first_half, first_half_sheet)
    _prepare_half(second_half, second_half_sheet)
    wb.save('testing1final.xls')
    wb.close()







