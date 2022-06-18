import openpyxl
from list_of_actions import times, teams
import datas



def create_template_of_excel(name, data_half_one, data_half_two):
    wb = openpyxl.Workbook()
    ws = wb.active
    _fill_halfs(ws, times)
    _fill_teams(ws, teams)
    _fill_actions(ws, datas.ACTIONS_FOR_EXCEL)
    fill_results(ws,data=data_half_one, start_col=2,check_list=datas.ACTIONS_FOR_EXCEL)
    fill_results(ws,data=data_half_two, start_col=4,check_list=datas.ACTIONS_FOR_EXCEL)
    wb.save(name)
    wb.close()


def _fill_halfs(ws: openpyxl.Workbook.active, halfs):
    col = 1
    for i in halfs:
        ws.cell(row=1, column=col, value=i)
        col += 3


def _fill_teams(ws: openpyxl.Workbook.active, teams):
    col = 2
    for team in teams:
        ws.cell(row=2, column=col, value=team)
        ws.cell(row=2, column=col + 2, value=team)
        col += 1


def _fill_actions(ws: openpyxl.Workbook.active, actions):
    row = 3
    for action in actions:
        ws.cell(row=row, column=1, value=action)
        row += 1

def fill_results(ws: openpyxl.Workbook.active, data, start_col,  check_list=datas.ACTIONS_FOR_EXCEL):

    col = start_col
    for team in data:
        row = 3
        for action in check_list:
            ws.cell(row=row, column=col, value=data[team][action])
            row += 1
        col += 1

