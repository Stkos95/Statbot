from openpyxl import  Workbook, load_workbook
import datas





def _get_result_activities_with_zones(data: dict, check_list, ws: Workbook.active):
    col_title = 4
    row_title = 11
    for zones in check_list:
        n = 0
        for activity in datas.ACTIONS_FOR_PERSONAL[::2]:
            ws.cell(row=row_title, column=col_title, value=data[zones][activity])
            n += 1
            col_title += 1
        col_title = col_title - n
        for activity in datas.ACTIONS_FOR_PERSONAL[1::2]:
            ws.cell(row=row_title + 1, column=col_title, value=data[zones][activity])
            col_title += 1


def _get_shoots_results(data:dict, ws: Workbook.active):
    col_title = 31
    row_title = 11
    for activity in datas.SHOOTS[::2]:
        ws.cell(row=row_title, column=col_title, value=data[activity])
        col_title += 1
    col_title = 31
    for activity in datas.SHOOTS[1::2]:
        ws.cell(row=row_title + 1, column=col_title, value=data[activity])
        col_title +=1

def _get_other_results(data: dict, ws: Workbook.active):
    row_title =  11
    col_title = 33
    for activity in datas.OTHER[:-2]:
        ws.cell(row=row_title, column=col_title, value=data[activity])
        col_title += 1
    for activity in datas.OTHER[-2:]:
        ws.cell(row=row_title, column=col_title, value=data[activity])
        row_title += 1



def _prepare_half(dict_data: dict, ws: Workbook.active):
    _get_result_activities_with_zones(dict_data,datas.ZONES[:3], ws)
    _get_shoots_results(dict_data['shoots'], ws)
    _get_other_results(dict_data['other'], ws)

def prepare_workbook(file_name,first_half, second_half):
    wb = load_workbook("Персональная статистика 2.0 - 4.xlsx")
    sh = wb.sheetnames[:2]
    first_half_sheet = wb[sh[0]]
    second_half_sheet = wb[sh[1]]
    _prepare_half(first_half, first_half_sheet)
    _prepare_half(second_half, second_half_sheet)
    wb.save(f'excel complete personal/{file_name}.xls')
    wb.close()