import openpyxl
from list_of_actions import times, teams, Teams
import datas

test_data = {'1 half': {'1': {'ЖК': 1,
                              'КК': 1,
                              'Обводка✅': 2,
                              'Обводка❌': 3,
                              'Угловые': 0,
                              'Фолы': 0,
                              'передача✅': 0,
                              'передача❌': 0,
                              'удар✅': 0,
                              'удар❌': 0},
                        '2': {'ЖК': 0,
                              'КК': 0,
                              'Обводка✅': 0,
                              'Обводка❌': 0,
                              'Угловые': 0,
                              'Фолы': 0,
                              'передача✅': 0,
                              'передача❌': 1,
                              'удар✅': 0,
                              'удар❌': 0}},
             '2 half': {'1': {'ЖК': 0,
                              'КК': 0,
                              'Обводка✅': 0,
                              'Обводка❌': 0,
                              'Угловые': 0,
                              'Фолы': 0,
                              'передача✅': 1,
                              'передача❌': 0,
                              'удар✅': 0,
                              'удар❌': 1},
                        '2': {'ЖК': 0,
                              'КК': 0,
                              'Обводка✅': 0,
                              'Обводка❌': 0,
                              'Угловые': 1,
                              'Фолы': 0,
                              'передача✅': 0,
                              'передача❌': 0,
                              'удар✅': 0,
                              'удар❌': 0}}}

test = Teams(team_one={'передача✅': 1, 'передача❌': 1, 'удар✅': 5, 'удар❌': 1, 'Обводка✅': 0, 'Обводка❌': 0, 'Сейвы': 0,
                       'Фолы': 0, 'ЖК': 0, 'КК': 0, 'Угловые': 0},
             team_two={'передача✅': 0, 'передача❌': 0, 'удар✅': 7, 'удар❌': 0, 'Обводка✅': 0, 'Обводка❌': 0, 'Сейвы': 0,
                       'Фолы': 0, 'ЖК': 0, 'КК': 50, 'Угловые': 1})


def transfer_results(results: Teams, check_list: list = None):
    passession_common = results.team_one['передача✅'] + results.team_two['передача✅']
    tmp = {}
    r = 1
    for team in results:
        print(team)
        shoots_all = sum([team[i] for i in team if i.startswith('удар')])
        try:
            percent_shoots = team['удар✅'] / shoots_all * 100
        except ZeroDivisionError:
            percent_shoots = 0
        pass_all = sum([team[i] for i in team if i.startswith('передача')])
        try:
            percent_pass = team['передача✅'] / pass_all * 100
        except ZeroDivisionError:
            percent_pass = 0
        obvodka_all = sum([team[i] for i in team if i.startswith('Обводка')])
        ball_posession = team['передача✅'] / passession_common
        tmp[r] = {"Удары всего": shoots_all,
                  "Удары точные": team['удар✅'],
                  "Процент точных ударов": percent_shoots,
                  "Передачи всего": pass_all,
                  "Передачи точные": team['передача✅'],
                  "Процент точных передач": percent_pass,
                  "Владение мячом": ball_posession,
                  "Обводка всего": obvodka_all,
                  "Обводка точные": team['Обводка✅'],
                  "Сейвы": team['Сейвы'],
                  "Фолы": team['Фолы'],
                  "ЖК": team['ЖК'],
                  "КК": team['КК'],
                  "Угловые": team['Угловые']

                  }
        r += 1

        print(tmp)


t = transfer_results(test)

("Удары всего",
 "Удары точные",
 "Процент точных ударов",
 "Передачи всего",
 "Передачи точные",
 "Процент точных передач",
 "Владение мячом",
 "Обводка всего",
 "Обводка точные",
 "Сейвы",
 "Фолы",
 "ЖК",
 "КК",
 "Угловые")


# def transfer_to_approriate_type(ws: openpyxl.Workbook.active,final_data: dict):
#     row = 3
#     col = 2
#     for half in final_data:
#         for team in final_data[half]:
#             row = 3
#
#             for action in final_data[half][team]:
#                 ws.cell(row=row, column=col, value=final_data[half][team][action])
#                 row+=1
#             col += 1


# def finalResults(results: dict[str]):
#     final_result = {}
#
#     for half in results:
#         final_result[half] = {}
#         for team in results[half]:
#             final_result[half][team] = {}
#             for action in results[half][team]:
#                 if not action.endswith("❌"):
#                     final_result[half][team][action] = results[half][team][action]
#                 elif action.endswith("❌"):
#                     tmpvalue = action.replace("❌", '✅')
#                     final_result[half][team][action.replace("❌","_всего")] = results[half][team][action] + results[half][team][tmpvalue]
#
#     return final_result


def create_template_of_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    _fill_halfs(ws, times)
    _fill_teams(ws, teams)
    _fill_actions(ws, datas.ACTIONS_FOR_EXCEL)
    # transfer_to_approriate_type(ws, test_data)
    wb.save('test1.xls')
    wb.close()


def _fill_halfs(ws: openpyxl.Workbook.active, halfs):
    col = 1
    for i in times:
        ws.cell(row=1, column=col, value=i)
        col += 3


def _fill_teams(ws: openpyxl.Workbook.active, teams):
    col = 2
    for team in teams:
        ws.cell(row=2, column=col, value=team)
        ws.cell(row=2, column=col + 2, value=team)
        col += 1


def _fill_actions(ws: openpyxl.Workbook.active, actions):
    row = 3
    for action in actions:
        ws.cell(row=row, column=1, value=action)
        row += 1


x = create_template_of_excel()
