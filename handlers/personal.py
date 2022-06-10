from loader import dp
from aiogram import types
from states import PersonalStatisticStates
from aiogram.dispatcher import FSMContext
from config import personal_template, Activities
from keyboards.kb_fabric import team_analitic_callback, personal_callback
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton
from pprint import pprint
from openpyxl import Workbook, load_workbook


def kb(data, forcd):
    kb = InlineKeyboardMarkup(row_width=2)
    for zone in data:
        if zone == forcd:
            for actions in data[zone]:
                quantity = data[zone][actions]

                d = InlineKeyboardButton(text=f"{actions} ({quantity})",
                                         callback_data=personal_callback.new(personal_action=f"{forcd}_{actions}"))
                kb.insert(d)
    return kb


@dp.callback_query_handler(text="personal", state=None)
async def personal_stat_start(call: types.CallbackQuery, state: FSMContext):
    await call.answer()

    await call.message.answer('Ниже 3 клавиатуры (разбиры по зонам) для подсчета первого тайма')
    async with state.proxy() as data:
        data["Первый тайм"] = personal_template.copy()

        for zone in data["Первый тайм"]:
            await call.message.answer(f"{zone}", reply_markup=kb(data["Первый тайм"], zone))
    await PersonalStatisticStates.First_state.set()


@dp.callback_query_handler(personal_callback.filter(), state=PersonalStatisticStates.First_state)
async def personal_test(call: types.CallbackQuery, state: FSMContext, callback_data: dict):
    await call.answer()
    print(callback_data)
    whatzone = callback_data['personal_action'].split("_")[0]
    whataction = callback_data['personal_action'].split("_")[1]
    async with state.proxy() as data:
        data["Первый тайм"][whatzone][whataction] += 1
        print(data["Первый тайм"][whatzone])
    await call.message.edit_reply_markup(reply_markup=kb(data["Первый тайм"], whatzone))


@dp.message_handler(text="Второй", state=PersonalStatisticStates.First_state)
async def start_second_half(message: types.Message, state: FSMContext):
    await message.answer("Вы закончили первый тайм и начали второй!!!")
    async with state.proxy() as data:
        data["Второй тайм"] = personal_template.copy()
        for zone in data["Второй тайм"]:
            await message.answer(f"{zone}-2 тайм", reply_markup=kb(data["Второй тайм"], zone))
    await PersonalStatisticStates.next()


@dp.callback_query_handler(personal_callback.filter(), state=PersonalStatisticStates.Second_state)
async def personal_test_1(call: types.CallbackQuery, state: FSMContext, callback_data: dict):
    await call.answer()
    print(callback_data)
    whatzone = callback_data['personal_action'].split("_")[0]
    whataction = callback_data['personal_action'].split("_")[1]
    async with state.proxy() as data:
        data["Второй тайм"][whatzone][whataction] += 1
        print(data["Второй тайм"][whatzone])
    await call.message.edit_reply_markup(reply_markup=kb(data["Второй тайм"], whatzone))


@dp.message_handler(text="Готово", state=PersonalStatisticStates.Second_state)
async def finishing(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        result = {
            "Первый тайм": data["Первый тайм"],
            "Второй тайм": data["Второй тайм"]
        }
    wb = load_workbook("Персональная статистика 2.0 - 4.xlsx")
    sheets = wb.sheetnames[:2]
    wb.active = wb[sheets[0]]
    ws = wb.active

    # for half in result:
    #     await message.answer(text=half)
    #     for zone in result[half]:
    #         answer = []
    #         for activity in result[half][zone]:
    #             answer.append(f"{activity} - {result[half][zone][activity]}")
    #             text = '\n'.join(answer)
    #         await message.answer(f"{zone}:\n{text}")
    # await state.finish()
    # pprint(result)

    for halfs in result:
        n = 1
        col_title = 4
        row_title = 11
        if halfs == "Второй тайм":
            wb.active = wb[sheets[1]]
            ws = wb.active

        for zones in result[halfs]:
            if zones != "Прочее":

                for activity in result[halfs][zones]:
                    if n % 2 != 0:
                        ws.cell(row=row_title, column=col_title, value=result[halfs][zones][activity])
                        n += 1
                    else:
                        ws.cell(row=row_title + 1, column=col_title, value=result[halfs][zones][activity])
                        n += 1
                        col_title += 1
            else:
                n = 1
                for activity in result[halfs][zones]:
                    if n != 6:
                        ws.cell(row=row_title, column=col_title, value=result[halfs][zones][activity])
                        n += 1

                    else:
                        ws.cell(row=row_title + 1, column=col_title, value=result[halfs][zones][activity])
                    col_title += 1


    # for halfs in result:
    #     ws.cell(row_title, col_title, value=halfs)
    #     for zones in result[halfs]:
    #         row_title += 1
    #         ws.cell(row=row_title, column=col_title, value=zones)
    #         for action in result[halfs][zones]:
    #             row_title += 1
    #             ws.cell(row=row_title, column=col_title + 1, value=action)
    #             ws.cell(row=row_title, column=col_title + 2, value=result[halfs][zones][action])
    #     row_title = 1
    #     col_title = 4

    wb.save('testing1.xls')
    wb.close()
